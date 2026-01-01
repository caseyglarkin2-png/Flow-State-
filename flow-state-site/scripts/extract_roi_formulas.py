#!/usr/bin/env python3

r"""Extract structure and formulas from an ROI .xlsx model.

Usage:
    python3 flow-state-site/scripts/extract_roi_formulas.py \
        --workbook "flow-state-site/roi-models/BT ROI CalculatorV2.xlsx"

Optional:
  --find "payback,total annual,detention,throughput,paper"   # keyword search
  --cells "Inputs!B2,Inputs!C3,Summary!H14"                   # explicit cell refs

Notes:
- This script prints:
  - sheet names
  - defined names / named ranges
  - formulas for any provided --cells
  - a lightweight keyword scan for formula cells

It does not attempt to fully reconstruct the model automatically; it’s designed
to quickly surface the mapping that we can then codify in docs/roi-model-spec.md.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from typing import Iterable, Optional, Tuple


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--workbook", required=True, help="Path to .xlsx workbook")
    p.add_argument(
        "--find",
        default="",
        help="Comma-separated keywords to search for in sheet values/formulas",
    )
    p.add_argument(
        "--cells",
        default="",
        help="Comma-separated cell refs like Sheet1!B2,Summary!H14",
    )
    p.add_argument(
        "--max-matches",
        type=int,
        default=80,
        help="Max keyword matches to print",
    )
    return p.parse_args()


@dataclass(frozen=True)
class CellRef:
    sheet: str
    addr: str


CELL_REF_RE = re.compile(r"^(?P<sheet>[^!]+)!(?P<addr>[A-Za-z]{1,3}[0-9]{1,7})$")


def parse_cell_refs(raw: str) -> list[CellRef]:
    raw = raw.strip()
    if not raw:
        return []

    out: list[CellRef] = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        m = CELL_REF_RE.match(part)
        if not m:
            raise ValueError(f"Invalid cell ref: {part} (expected Sheet!A1)")
        out.append(CellRef(sheet=m.group("sheet"), addr=m.group("addr").upper()))
    return out


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def main() -> int:
    args = parse_args()

    try:
        from openpyxl import load_workbook
    except Exception as e:  # pragma: no cover
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        print(f"Import error: {e}")
        return 2

    wb = load_workbook(args.workbook, data_only=False)

    print(f"Workbook: {args.workbook}")
    print("\nSheets:")
    for name in wb.sheetnames:
        print(f"- {name}")

    print("\nDefined names (named ranges):")
    # openpyxl exposes defined names via a dict-like object in 3.1.x
    defined_items = list(getattr(wb, "defined_names", {}).items())
    if not defined_items:
        print("- (none)")
    else:
        for name, dn in defined_items:
            # dn.attr_text contains the formula/range like 'Sheet1!$A$1:$B$2'
            attr_text = getattr(dn, "attr_text", "")
            print(f"- {name}: {attr_text}")

    cell_refs = parse_cell_refs(args.cells)
    if cell_refs:
        print("\nExplicit cells:")
        for ref in cell_refs:
            if ref.sheet not in wb.sheetnames:
                print(f"- {ref.sheet}!{ref.addr}: (missing sheet)")
                continue
            ws = wb[ref.sheet]
            cell = ws[ref.addr]
            value = cell.value
            is_formula = isinstance(value, str) and value.startswith("=")
            kind = "FORMULA" if is_formula else "VALUE"
            print(f"- {ref.sheet}!{ref.addr} [{kind}]: {_cell_str(value)}")

    keywords = [k.strip().lower() for k in args.find.split(",") if k.strip()]
    if keywords:
        print("\nKeyword scan (matches values or formulas):")
        matches = 0
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    raw = cell.value
                    if raw is None:
                        continue
                    s = _cell_str(raw).lower()
                    if any(k in s for k in keywords):
                        addr = cell.coordinate
                        is_formula = isinstance(raw, str) and raw.startswith("=")
                        kind = "FORMULA" if is_formula else "VALUE"
                        print(f"- {sheet}!{addr} [{kind}]: {_cell_str(raw)}")
                        matches += 1
                        if matches >= args.max_matches:
                            print(f"(stopping at {args.max_matches} matches)")
                            return 0

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
